# -*- coding: utf-8 -*-
"""Нейтральные документы-образцы для снимков экрана в инструкции.

Никаких настоящих рабочих файлов: всё содержимое выдумано и обезличено, чтобы
картинки в инструкции можно было показывать кому угодно.

Язык образцов задаётся ключом --lang: в английском руководстве русские имена файлов
на снимках выглядели бы так же неуместно, как английские в русском.

Usage: python make_samples.py [--lang ru|en]
"""
import argparse
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

PARSER = argparse.ArgumentParser()
PARSER.add_argument("--lang", choices=("ru", "en"), default="ru")
PARSER.add_argument("--out", default=None, help="куда складывать образцы")
ARGS = PARSER.parse_args()
LANG = ARGS.lang

OUT = os.path.dirname(os.path.abspath(__file__))
SAMPLES = ARGS.out or os.path.join(OUT, "samples-" + LANG)
os.makedirs(SAMPLES, exist_ok=True)

FONTS = r"C:\Windows\Fonts"
pdfmetrics.registerFont(TTFont("TNR", os.path.join(FONTS, "times.ttf")))
pdfmetrics.registerFont(TTFont("TNR-Bold", os.path.join(FONTS, "timesbd.ttf")))

# Весь видимый текст образцов — здесь: подставлять его по месту значило бы держать
# два почти одинаковых скрипта, и они разошлись бы при первой же правке.
TEXT = {
    "ru": {
        "body": [
            "Настоящий документ подготовлен в качестве образца для демонстрации "
            "возможностей программы. Любое совпадение с реальными документами случайно.",
            "Раздел описывает порядок действий при подготовке комплекта материалов: "
            "сбор исходных файлов, проверку состава, формирование итогового документа "
            "и передачу результата ответственному исполнителю.",
            "Исполнитель проверяет комплектность материалов, наличие подписей и "
            "приложений. При обнаружении расхождений материалы возвращаются на "
            "доработку с указанием причин.",
            "Результат оформляется отдельным файлом. Исходные материалы сохраняются "
            "в неизменном виде и могут быть использованы повторно.",
        ],
        "table": [("№", "Наименование", "Кол-во", "Примечание"),
                  ("1", "Пояснительная записка", "1", "на 4 листах"),
                  ("2", "Расчёт показателей", "1", "приложение А"),
                  ("3", "Сводная таблица", "2", "приложение Б"),
                  ("4", "Копии исходных материалов", "5", "—")],
        "pageOf": "%d из %d",
        "doc": "Образец документа",
        "docSections": ["Раздел 1. Общие положения", "Раздел 2. Порядок подготовки",
                        "Раздел 3. Состав приложений", "Раздел 4. Проверка комплектности",
                        "Раздел 5. Оформление результата", "Раздел 6. Передача материалов",
                        "Раздел 7. Хранение", "Раздел 8. Заключительные положения"],
        "appA": "Приложение А", "appB": "Приложение Б", "sheet": "Лист",
        "xlDir": "Книги Excel", "unit": "Отдел",
        "xlTitle": "Показатели подразделения «%s»",
        "xlHead": ["№", "Показатель", "План", "Факт"],
        "xlRow": "Показатель %d", "xlTotal": "Итого",
        "photo": "Снимок страницы", "photoTitle": "СНИМОК СТРАНИЦЫ",
        "photoLine": "Строка %d текста на снятой странице",
        "sayPdf": "Образцы PDF:", "sayXl": "Образцы Excel:", "sayDone": "Готово:", "pages": "стр.",
        "sayImg": "Образец снимка:",
    },
    "en": {
        "body": [
            "This document has been prepared as a sample, to show what the program does. "
            "Any resemblance to a real document is coincidental.",
            "The section sets out the order of work when a set of materials is put "
            "together: collecting the source files, checking that the set is complete, "
            "producing the final document and handing the result to the person responsible.",
            "The person responsible checks that the set is complete and that the signatures "
            "and appendices are in place. Where something does not agree, the materials go "
            "back for revision with the reasons given.",
            "The result is issued as a separate file. The source materials are kept "
            "unchanged and may be used again.",
        ],
        "table": [("No.", "Item", "Qty", "Note"),
                  ("1", "Explanatory note", "1", "4 sheets"),
                  ("2", "Calculation of indicators", "1", "appendix A"),
                  ("3", "Summary table", "2", "appendix B"),
                  ("4", "Copies of source materials", "5", "—")],
        "pageOf": "%d of %d",
        "doc": "Sample document",
        "docSections": ["Section 1. General provisions", "Section 2. Order of preparation",
                        "Section 3. Contents of the appendices", "Section 4. Checking the set",
                        "Section 5. Issuing the result", "Section 6. Handing over",
                        "Section 7. Retention", "Section 8. Final provisions"],
        "appA": "Appendix A", "appB": "Appendix B", "sheet": "Sheet",
        "xlDir": "Excel workbooks", "unit": "Division",
        "xlTitle": "Indicators of the %s division",
        "xlHead": ["No.", "Indicator", "Plan", "Actual"],
        "xlRow": "Indicator %d", "xlTotal": "Total",
        "photo": "Page photo", "photoTitle": "PAGE PHOTO",
        "photoLine": "Line %d of text on the photographed page",
        "sayPdf": "Sample PDFs:", "sayXl": "Sample workbooks:", "sayDone": "Done:", "pages": "pages",
        "sayImg": "Sample photo:",
    },
}
T = TEXT[LANG]

BODY = T["body"]


def page(c, title, number, total, table=False):
    w, h = A4
    c.setFont("TNR-Bold", 16)
    c.drawString(25 * mm, h - 30 * mm, title)
    c.setLineWidth(0.7)
    c.line(25 * mm, h - 33 * mm, w - 20 * mm, h - 33 * mm)

    y = h - 45 * mm
    c.setFont("TNR", 12)
    for para in BODY:
        words, line = para.split(), ""
        for word in words:
            probe = (line + " " + word).strip()
            if c.stringWidth(probe, "TNR", 12) > w - 45 * mm:
                c.drawString(25 * mm, y, line)
                y -= 6.5 * mm
                line = word
            else:
                line = probe
        if line:
            c.drawString(25 * mm, y, line)
            y -= 10 * mm

    if table:
        rows = T["table"]
        col = [15 * mm, 75 * mm, 25 * mm, 45 * mm]
        x0 = 25 * mm
        for r, row in enumerate(rows):
            x = x0
            c.setFont("TNR-Bold" if r == 0 else "TNR", 11)
            for i, cell in enumerate(row):
                c.rect(x, y - 8 * mm, col[i], 8 * mm)
                c.drawString(x + 2 * mm, y - 5.5 * mm, cell)
                x += col[i]
            y -= 8 * mm

    c.setFont("TNR", 10)
    c.drawCentredString(w / 2, 15 * mm, T["pageOf"] % (number, total))
    c.showPage()


def build(path, titles, bookmarks=False):
    c = canvas.Canvas(path, pagesize=A4)
    total = len(titles)
    for i, title in enumerate(titles, start=1):
        if bookmarks:
            key = "sec%d" % i
            c.bookmarkPage(key)
            c.addOutlineEntry(title, key, level=0)
        page(c, title, i, total, table=(i % 3 == 0))
    c.save()
    print("  ", os.path.basename(path), "-", total, T["pages"])


print(T["sayPdf"])
build(os.path.join(SAMPLES, T["doc"] + ".pdf"), T["docSections"], bookmarks=True)
build(os.path.join(SAMPLES, T["appA"] + ".pdf"),
      ["%s. %s %d" % (T["appA"], T["sheet"], i) for i in (1, 2, 3, 4)])
build(os.path.join(SAMPLES, T["appB"] + ".pdf"),
      ["%s. %s %d" % (T["appB"], T["sheet"], i) for i in (1, 2, 3)])

# Книги Excel для экрана «Свод Excel».
XL = os.path.join(SAMPLES, T["xlDir"])
os.makedirs(XL, exist_ok=True)
print(T["sayXl"])
for n, name in enumerate(["%s %d" % (T["unit"], i) for i in range(1, 6)], start=1):
    wb = Workbook()
    ws = wb.active
    ws.title = name
    ws["A1"] = T["xlTitle"] % name
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")
    head = T["xlHead"]
    for col, text in enumerate(head, start=1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    for row in range(4, 12):
        ws.cell(row=row, column=1, value=row - 3)
        ws.cell(row=row, column=2, value=T["xlRow"] % (row - 3))
        ws.cell(row=row, column=3, value=(row - 3) * 10 * n)
        ws.cell(row=row, column=4, value=(row - 3) * 9 * n)
    ws.cell(row=12, column=2, value=T["xlTotal"])
    ws.cell(row=12, column=3, value="=SUM(C4:C11)")
    ws.cell(row=12, column=4, value="=SUM(D4:D11)")
    for col, width in zip("ABCD", (6, 34, 12, 12)):
        ws.column_dimensions[col].width = width
    path = os.path.join(XL, "%s.xlsx" % name)
    wb.save(path)
    print("  ", os.path.basename(path))

# Образец снимка — то, что человек добавляет в «Прочих операциях» кнопкой «Добавить
# картинки…»: лист, снятый под небольшим углом на сероватом фоне. Рисуем сами, потому что
# ни одна настоящая фотография не должна попасть ни в репозиторий, ни в руководство.
print(T["sayImg"])
photo = os.path.join(SAMPLES, T["photo"] + ".jpg")
try:
    from PIL import Image, ImageDraw, ImageFont

    img = Image.new("RGB", (1400, 1000), (168, 170, 174))
    d = ImageDraw.Draw(img)
    d.rectangle([120, 60, 1280, 940], fill=(250, 249, 246), outline=(196, 196, 196))
    try:
        head = ImageFont.truetype("times.ttf", 44)
        body = ImageFont.truetype("times.ttf", 26)
    except OSError:
        head = body = ImageFont.load_default()
    d.text((200, 120), T["photoTitle"], font=head, fill=(30, 30, 30))
    for i in range(1, 17):
        d.text((200, 210 + i * 42), T["photoLine"] % i, font=body, fill=(48, 48, 48))
    img.save(photo, "JPEG", quality=88)
    print("  ", os.path.basename(photo))
except ImportError:
    print("   PIL не установлен — снимок не создан")

print(T["sayDone"], SAMPLES)
